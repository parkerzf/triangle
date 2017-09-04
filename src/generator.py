import sys, os
import numpy as np
from sets import Set
import pickle as pkl
import networkx as nx
from openpyxl import load_workbook

def gen_num_trips(is_skew):
	if is_skew:
		table = [0.7, 0.2, 0.1]
		return np.random.choice(np.arange(2, 5), p=table)
	return np.random.choice(np.arange(2, 5))

def gen_prob_table(num_hotspots, start_points):
	num_start_points = len(start_points)
	start_points_list = start_points.tolist()
	if num_hotspots is 0:
		table = [0.7/num_start_points] * (num_start_points+1)
	else:
		set = Set([])
		hotspots_prob, others_prob = calc_prob(num_start_points, num_hotspots)

		for i in range(num_hotspots):
			a = np.random.choice(start_points)
			while a in set:
				a = np.random.choice(start_points)
			set.add(a)

		table = [others_prob] * (num_start_points+1)
		for a in set:
			i = start_points_list.index(a)
			table[i+1] = hotspots_prob

	table[0] = 0.3

	return table

def calc_prob(num_start_points, num_hotspots):
	others_prob = 7.7/(150-14*num_hotspots + 11*num_start_points)
	hotspots_prob = 3*(50-num_hotspots) * others_prob / (11*num_hotspots)

	return hotspots_prob, others_prob



def gen_start_prob_table(num_hotspots, start_points):
	num_start_points = len(start_points)
	start_points_list = start_points.tolist()
	if num_hotspots is 0:
		table = [1.0/num_start_points] * num_start_points
	else:
		set = Set([])

		hotspots_prob, others_prob = calc_prob(num_start_points, num_hotspots)
		hotspots_prob = hotspots_prob * 10 / 7
		others_prob = others_prob * 10 / 7

		for i in range(num_hotspots):
			a = np.random.choice(start_points)
			while a in set:
				a = np.random.choice(start_points)
			set.add(a)

		table = [others_prob] * (num_start_points)
		for a in set:
			i = start_points_list.index(a)
			table[i] = hotspots_prob

	return table

def random_choice_start_points(num_start_points=50):
	return np.random.choice(np.arange(1, 51), num_start_points, replace=False)

def gen_seq_points(num_trips, start_table, table, start_points):
	seq_points = [0] * num_trips
	set = Set([])

	seq_points[0] = np.random.choice(start_points, p = start_table)
	# seq_points[0] = np.random.choice(start_points)
	set.add(seq_points[0])

	middle_points = np.insert(start_points, 0, 0)

	for i in range(1, num_trips):
		a = np.random.choice(middle_points, p = table)
		while a in set:
			a = np.random.choice(middle_points, p = table)
		set.add(a)
		seq_points[i] = a

	return seq_points

def gen_rand_between(start, end):
	return np.random.choice(np.arange(start, end + 1))


def gen_sp_path(user_id, seq_points, sp_matrix, ssp, lsp, graph):
	num_trips = len(seq_points)
	seq_points.append(seq_points[0])
	format_str = ''
	format_full_str = ''
	seq_latest_times = [0] * num_trips
	seq_earliest_times = [0] * num_trips
	
	for i in range(num_trips):
		shortest_path_len = int(round(sp_matrix[str(seq_points[i])][str(seq_points[i+1])]))
		# slack = int(round(scale(sp_matrix[str(seq_points[i])][str(seq_points[i+1])], ssp, lsp, 20, 90)))
		slack = 30
		if i == 0:
			seq_latest_times[0] = gen_rand_between(480, 540)
			seq_earliest_times[0] = seq_latest_times[0] - shortest_path_len - slack
		elif i == num_trips -1:
			if seq_latest_times[i-1] > 960:
				seq_earliest_times[num_trips-1] = seq_latest_times[i-1] + 60
			else:
				seq_earliest_times[num_trips-1] = gen_rand_between(1020, 1140)
			seq_latest_times[num_trips-1] = seq_earliest_times[num_trips-1] + shortest_path_len + slack
		else:
			earliest_latest = seq_latest_times[i-1] + shortest_path_len + slack + 60
			if earliest_latest > 960:
				seq_latest_times[i] = earliest_latest
			else:
				seq_latest_times[i] = gen_rand_between(earliest_latest, 960)
			seq_earliest_times[i] = seq_latest_times[i] - shortest_path_len - slack

	for i in range(num_trips):
		format_str += str(user_id+1) + '\t' + str(i+1) + '\t' + \
		str(seq_points[i]) + '\t' + str(seq_points[i+1]) + \
			'\t' + str(seq_latest_times[i]) + '\t'+ str(3) + '\t' + str(seq_earliest_times[i]) + '\n'

		shortest_path = nx.shortest_path(graph, str(seq_points[i]), str(seq_points[i+1]), weight = "weight")
		format_full_str += gen_full_sp_path(user_id, i, shortest_path, graph)

	return format_str, format_full_str

def gen_full_sp_path(user_id, trip_id, shortest_path, graph):
	format_str = ''
	for i in range(len(shortest_path)-1):
		s = str(shortest_path[i])
		t = str(shortest_path[i+1])
		weight = graph[s][t]['weight']
		format_str += str(user_id+1) + '\t' + str(trip_id+1) + '\t' + \
		s + '\t' + t + '\t' \
		+ str(1) + '\t'+ str(weight) + '\n'

	return format_str



def scale(valueIn, baseMin, baseMax, limitMin, limitMax):
    return ((limitMax - limitMin) * (valueIn - baseMin) / (baseMax - baseMin)) + limitMin


def clear_ws(ws):
	for row in ws.iter_rows(min_row=2):
		for cell in row:
			cell.value=None

def write_summary_excel(ws, idx_summary, summary_str):
	for line in summary_str.split('\n'):
		elems = line.split('\t')
		if len(elems) == 7:
			for i, elem in enumerate(elems):
				# 73 is ASCII code of I
				ws['%c%d' % (73 + i, idx_summary)] = float(elems[i])
			idx_summary += 1
	return idx_summary

def write_full_excel(ws, idx_full, full_str):
	for line in full_str.split('\n'):
		elems = line.split('\t')
		if len(elems) == 6:
			for i, elem in enumerate(elems):
				# 65 is ASCII code of A
				ws['%c%d' % (65 + i, idx_full)] = float(elems[i])
			idx_full += 1
	return idx_full

graph_path = sys.argv[1]
dir_name = sys.argv[2]
num_user = int(sys.argv[3])
num_hotspots = int(sys.argv[4])
num_start_points = int(sys.argv[5])
seq = int(sys.argv[6])
is_skew = True if int(sys.argv[7]) == 1 else False
excel_template = sys.argv[8]

wb = load_workbook(excel_template, keep_vba=True)
ws = wb['Model Input']

path_graph = nx.read_graphml(graph_path)
sp_matrix = nx.shortest_path_length(path_graph, weight = "weight")

# prefix = graph_path[0: graph_path.find('.', graph_path.find('.') + 1)]
prefix = os.path.splitext(os.path.basename(graph_path))[0]

directory = '{}/{}_{}_{}_{}'.format(dir_name, prefix, num_user, num_hotspots, num_start_points)
if not os.path.exists(directory):
    os.makedirs(directory)

ssp = float("inf")
lsp = 0

for i in range(0, len(sp_matrix)):
	for j in range(i+1, len(sp_matrix)):
		if sp_matrix[str(i)][str(j)] < ssp:
			ssp = sp_matrix[str(i)][str(j)]
		if sp_matrix[str(i)][str(j)] > lsp:
			lsp = sp_matrix[str(i)][str(j)]

for i in range(1, seq+1):
	start_points = random_choice_start_points(num_start_points)

	table = gen_prob_table(num_hotspots, start_points)
	start_table = gen_start_prob_table(num_hotspots, start_points)
	# Open a file
	f_summary = open('{}/{}_{}_{}_{}_{}.txt'.format(directory, prefix, num_user, num_hotspots, num_start_points, i), "wb")
	f_full = open('{}/{}_{}_{}_{}_{}_full.txt'.format(directory, prefix, num_user, num_hotspots, num_start_points, i), "wb")

	idx_summary = 2
	idx_full = 2
	for j in range(0, num_user):
		num_trips = gen_num_trips(is_skew)
		seq_points = gen_seq_points(num_trips, start_table, table, start_points)
		summary_str, full_str = gen_sp_path(j, seq_points, sp_matrix, ssp, lsp, path_graph)
		if summary_str != '':
			f_summary.write(summary_str)
			idx_summary = write_summary_excel(ws, idx_summary, summary_str)

			f_full.write(full_str)
			idx_full = write_full_excel(ws, idx_full, full_str)

	# Close opend file
	f_summary.close()
	f_full.close()

	excel_path = '{}/{}_{}_{}_{}_{}.xlsm'.format(directory, prefix, num_user, num_hotspots, num_start_points, i)
	wb.save(excel_path)
	clear_ws(ws)
